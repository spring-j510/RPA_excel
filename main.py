from openpyxl import Workbook
import function



class candidate:
    def __init__(self, name,mail,mobile,motive):
        self.name = name
        self.mail = mail
        self.mobile = mobile
        self.motive = motive

    def print_tot(self):
        tot = self.name+self.mobile+self.motive
        print(tot)

myinstance = candidate(1,2,3,4)
myinstance.print_tot()

print(candidate.name)

"""
name:
mail:
mobile:
adress:
reason for (one's) desire:
"""


# # ワークブックの新規作成と保存
# wb = Workbook()
# wb.save(function.excel_title())

# # ワークブックの読み込み
# from openpyxl import load_workbook

# wb = load_workbook(function.excel_title())

# # 変更したいシート名を指定する
# old_sheet_name = 'Sheet'
# new_sheet_name = 'summary'

# # シート名を変更する
# ws = wb[old_sheet_name]
# ws.title = new_sheet_name


# # ワークシートの選択
# ws = wb['summary']  # ワークシートを指定
# ws = wb.active  # アクティブなワークシートを選択

# # セルに書き込み
# ws['B2'] = 'No.'
# ws['C2'] = 'name'
# ws['D2'] = 'Mail'
# ws['E2'] = 'Mobile'
# ws['F2'] = 'Adress'
# ws['G2'] = 'reason for (ones) desire'







# wb.save(function.excel_title())  # overwrite myworkbook.xlsx